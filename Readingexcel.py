import openpyxl


file = "C:\\Users\\Shrisha\\OneDrive\\data.xlsx"

wb=openpyxl.load_workbook(file)
# print(type(wb))

sheet = wb['Sheet2']

# row = sheet.max_row
# print(row)
# col = sheet.max_column
# print(col)
#
# for r in range(1,row+1):
#     for c in range(1,col+1):
#         print(sheet.cell(r,c).value, end='         ')
#     print()

data = [(1,2),(3,4),(5,6)]

for r in range(1,4):
    for c in range(1,4):
        sheet.cell(r,c).value = data[r]

wb.save(file)
